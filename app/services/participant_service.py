"""Import participants from Excel files."""
import logging
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.participant import Participant

logger = logging.getLogger(__name__)

# Expected column headers (case-insensitive matching)
COLUMN_MAP = {
    "dossard": "bib_number",
    "bib": "bib_number",
    "bib_number": "bib_number",
    "numero": "bib_number",
    "nom": "last_name",
    "last_name": "last_name",
    "name": "last_name",
    "prenom": "first_name",
    "prénom": "first_name",
    "first_name": "first_name",
    "firstname": "first_name",
    "email": "email",
    "mail": "email",
    "temps": "race_time",
    "time": "race_time",
    "race_time": "race_time",
    "chrono": "race_time",
    "pays": "country",
    "country": "country",
    "nationalite": "country",
    "nationalité": "country",
}


def import_participants_from_excel(
    db: Session, event_id: int, file_bytes: bytes
) -> dict:
    """Parse an Excel file and import participants into the database.

    Returns dict with imported, skipped, errors counts.
    """
    wb = load_workbook(filename=BytesIO(file_bytes), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"imported": 0, "skipped": 0, "errors": ["Fichier vide"]}

    # Map header row to field names
    header = rows[0]
    col_mapping = {}
    for idx, cell in enumerate(header):
        if cell is None:
            continue
        key = str(cell).strip().lower().replace(" ", "_")
        if key in COLUMN_MAP:
            col_mapping[idx] = COLUMN_MAP[key]

    if "bib_number" not in col_mapping.values():
        return {
            "imported": 0,
            "skipped": 0,
            "errors": ["Colonne dossard/bib introuvable dans l'en-tete"],
        }

    # Delete existing participants for this event (replace mode)
    db.query(Participant).filter(Participant.event_id == event_id).delete()

    imported = 0
    skipped = 0
    errors = []

    for row_idx, row in enumerate(rows[1:], start=2):
        try:
            data = {}
            for col_idx, field_name in col_mapping.items():
                val = row[col_idx] if col_idx < len(row) else None
                if val is not None:
                    data[field_name] = str(val).strip()

            bib = data.get("bib_number", "").strip()
            if not bib:
                skipped += 1
                continue

            participant = Participant(
                event_id=event_id,
                bib_number=bib,
                first_name=data.get("first_name"),
                last_name=data.get("last_name"),
                email=data.get("email"),
                race_time=data.get("race_time"),
                country=data.get("country"),
            )
            db.add(participant)
            imported += 1
        except Exception as e:
            errors.append(f"Ligne {row_idx}: {str(e)}")

    db.commit()
    wb.close()

    return {"imported": imported, "skipped": skipped, "errors": errors}
